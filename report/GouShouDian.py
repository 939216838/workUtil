import datetime
import os

from openpyxl.cell import Cell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

global global_self


def is_none(cell):
    if cell.value is None:
        print(f"当前单元格{cell.row}行{cell.column}列的值为空")
        cell: Cell
        cell.value = float(0)


def check_none(*cells):
    for cell in cells:
        if cell.value is None:
            print(f"当前单元格{cell.row}行{cell.column}列的值为空")
            cell.value = float(0)


# 点击了开始按钮
def start(self, path, wx):
    # 获取指定path路径下的文件列表
    list_file_name = get_list_file_by_path(wx, path)
    if list_file_name == -1:
        set_m_gauge_value(self, 0)
        return

    diyibu(path, list_file_name)
    dierbu(path, list_file_name)
    disanbu(path, list_file_name)
    disibu(path, list_file_name)
    diwubu(path, list_file_name)
    laststep(path, list_file_name)

    print("工作结束")
    set_m_gauge_value(self, 100)


# 第一步(本年累计)填趸售用电（对系统外趸售企业售电）--地方水电电量 电费
def diyibu(path, list_file_name):
    # 获取workbook对象,目的是用来保存当前excel表格.   获取sheet对象,目的是操作某一sheet工作表
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "计算过程并表", "本年累计", True,
                                                            True)

    cell1 = sheet.cell(75, 3)
    cell2 = sheet.cell(75, 41)
    if cell2.value is None:
        print("是none 执行这里,为了代码不报错,后续对这个值有操作,给此空值赋值为0")
        cell2.value = 0.0

    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)
    sheet.cell(53, 5).value = cell1.value
    sheet.cell(107, 5).value = cell2.value * 1000

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")

    workbook.save(route)


# 第二步(本年累计)填用于省内直接参与市场用户数据,即延吉天楹市场化电量 电费
def dierbu(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "外购", "2023年", True,
                                                            True)
    cell1 = sheet.cell(45, 9)
    if cell1.value is None:
        print("是none 执行这里,为了代码不报错,后续对这个值有操作,给此空值赋值为0")
        cell1.value = 0.0
    cell2 = sheet.cell(45, 11)
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)

    sheet.cell(40, 5).value = cell1.value / 10
    sheet.cell(94, 5).value = cell2.value

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")

    workbook.save(route)


# 第三步(本年累计)填两部制大工业用户基本电费
def disanbu(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "计算过程全口径", "本年累计", True,
                                                            True)
    cell1 = sheet.cell(29, 25)
    cell2 = sheet.cell(29, 26)
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)

    sheet.cell(100, 5).value = cell1.value + cell2.value

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")

    workbook.save(route)


# 第四步(上年同期累计)填用于省内直接参与市场用户数据,即延吉天楹市场化电量 电费
def disibu(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "天楹2022年", "Sheet1", True,
                                                            True)
    # 获取当前月份
    month = str(datetime.datetime.now().month % 12)

    # 定义sheet对象,有代码提示
    sheet: Worksheet

    # sheet.max_row最大行,但是for循环 for row in max_row ,不包含max_row,例   1 in 5   实际输出1234,所以+1
    max_row = sheet.max_row + 1

    # 当年累计数量
    累计数量 = 0.0
    累计电费 = 0.0
    # 数字的for循环,  需要用range(max_row),对象的for循环可以直接用in,从2开始到 max_row,但是不包含max_row所以+1
    for row in range(2, max_row):

        cell: EmptyCell = sheet.cell(row, 1)
        print("当前行是 ", row, "当前值是: ", cell.value)
        var = sheet.cell(row, 1).value
        # 从var里获取月份
        当前月份 = get_numbers(str(var))

        if 当前月份 <= month and (str(sheet.cell(row, 3).value) != "0.3731"):
            pw_num = sheet.cell(row, 3).value
            print("当前的电价是:", pw_num)
            累计数量 = float(sheet.cell(row, 2).value) + 累计数量
            累计电费 = float(sheet.cell(row, 6).value) + 累计电费
        elif 当前月份 > month:
            break

    print("最终的值是", 累计数量)
    print("最终的电费值是", 累计电费)

    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)

    sheet.cell(40, 6).value = 累计数量 / 10000
    sheet.cell(94, 6).value = 累计电费

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")

    workbook.save(route)


# 第五步(上年同期累计)填两部制大工业用户基本电费
def diwubu(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "电力销售月报表", "Sheet", True,
                                                            True)
    cell1 = sheet.cell(29, 25)
    cell2 = sheet.cell(29, 26)
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)

    sheet.cell(100, 6).value = cell1.value + cell2.value

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")

    workbook.save(route)


# 最后一步进行表内计算
def laststep(path, list_file_name):
    sheet, manual_table_name, workbook = get_workbook_sheet(path, list_file_name, "空白表样", "购售电情况表", False,
                                                            False)

    # 第一种 获得cell
    # 定义变量类型的方式之一  变量名: 类名
    sheet: Worksheet

    # 开始我的表格计算
    # 本年累计数
    cell_e21 = sheet.cell(21, 5)
    cell_e40 = sheet.cell(40, 5)
    check_none(cell_e21, cell_e40)
    value1 = cell_e21.value
    value2 = cell_e40.value
    sheet.cell(42, 5, float(value1) - float(value2))

    cell_e46 = sheet.cell(46, 5)
    cell_e47 = sheet.cell(47, 5)
    check_none(cell_e46, cell_e47)
    value1 = cell_e46.value
    value2 = cell_e47.value
    sheet.cell(63, 5, (float(value1) + float(value2)) * float(0.5376))

    cell_e53 = sheet.cell(53, 5)
    cell_e55 = sheet.cell(55, 5)
    check_none(cell_e53, cell_e55)
    value3 = cell_e53.value
    value4 = cell_e55.value
    sheet.cell(64, 5, (float(value1) + float(value2)) * float(0.4624) + float(value3) + float(value4))

    cell_e51 = sheet.cell(51, 5)
    cell_e52 = sheet.cell(52, 5)
    check_none(cell_e51, cell_e52)
    value1 = cell_e51.value
    value2 = cell_e52.value
    sheet.cell(65, 5, float(value1) + float(value2))

    cell_e92 = sheet.cell(92, 5)
    cell_e94 = sheet.cell(94, 5)
    check_none(cell_e92, cell_e94)
    value1 = cell_e92.value
    value2 = cell_e94.value
    sheet.cell(96, 5, float(value1) - float(value2))

    cell_e99 = sheet.cell(99, 5)
    cell_e101 = sheet.cell(101, 5)
    check_none(cell_e99, cell_e101)
    value1 = cell_e99.value
    value2 = cell_e101.value
    sheet.cell(117, 5, (float(value1) + float(value2)) * float(0.5006))

    cell_e107 = sheet.cell(107, 5)
    cell_e109 = sheet.cell(109, 5)
    check_none(cell_e107, cell_e109)
    value3 = cell_e107.value
    value4 = cell_e109.value
    sheet.cell(119, 5, (float(value1) + float(value2)) * float(0.4994) + float(value3) + float(value4))

    cell_e105 = sheet.cell(105, 5)
    cell_e106 = sheet.cell(106, 5)
    check_none(cell_e105, cell_e106)
    value1 = cell_e105.value
    value2 = cell_e106.value
    sheet.cell(121, 5, float(value1) + float(value2))

    cell_e125 = sheet.cell(125, 5)
    cell_e126 = sheet.cell(126, 5)
    check_none(cell_e126,cell_e125)
    cell_e126.value = cell_e125.value

    # 上年同期累计数
    cell_f21 = sheet.cell(21, 6)
    cell_f40 = sheet.cell(40, 6)
    check_none(cell_f21, cell_f40)
    value1 = cell_f21.value
    value2 = cell_f40.value
    # 上年同期累计数
    sheet.cell(42, 6, float(value1) - float(value2))


    cell_f46 = sheet.cell(46, 6)
    cell_f47 = sheet.cell(47, 6)
    check_none(cell_f46, cell_f47)
    value1 = cell_f46.value
    value2 = cell_f47.value
    sheet.cell(63, 6, (float(value1) + float(value2)) * float(0.5376))

    cell_f53 = sheet.cell(53, 6)
    cell_f55 = sheet.cell(55, 6)
    check_none(cell_f53, cell_f55)
    value3 = cell_f53.value
    value4 = cell_f55.value
    sheet.cell(64, 6, (float(value1) + float(value2)) * float(0.4624) + float(value3) + float(value4))

    cell_f51 = sheet.cell(51, 6)
    cell_f52 = sheet.cell(52, 6)
    check_none(cell_f51, cell_f52)
    value1 = cell_f51.value
    value2 = cell_f52.value
    sheet.cell(65, 6, float(value1) + float(value2))

    cell_f92 = sheet.cell(92, 6)
    cell_f94 = sheet.cell(94, 6)
    check_none(cell_f92, cell_f94)
    value1 = cell_f92.value
    value2 = cell_f94.value
    sheet.cell(96, 6, float(value1) - float(value2))

    cell_f99 = sheet.cell(99, 6)
    cell_f101 = sheet.cell(101, 6)
    check_none(cell_f99, cell_f101)
    value1 = cell_f99.value
    value2 = cell_f101.value
    sheet.cell(117, 6, (float(value1) + float(value2)) * float(0.5006))

    cell_f107 = sheet.cell(107, 6)
    cell_f109 = sheet.cell(109, 6)
    check_none(cell_f107, cell_f109)
    value3 = cell_f107.value
    value4 = cell_f109.value
    sheet.cell(119, 6, (float(value1) + float(value2)) * float(0.4994) + float(value3) + float(value4))

    cell_f105 = sheet.cell(105, 6)
    cell_f106 = sheet.cell(106, 6)
    check_none(cell_f105, cell_f106)
    value1 = cell_f105.value
    value2 = cell_f106.value
    sheet.cell(121, 6, float(value1) + float(value2))

    cell_f125 = sheet.cell(125, 6)
    cell_f126 = sheet.cell(126, 6)
    check_none(cell_f126,cell_f125)
    cell_f126.value = cell_f125.value

    route, manual_table_name = get_file_path(path, list_file_name, "空白表样")
    workbook.save(route)
    pass

# 封装 根据文件名,sheet名,获取表格操作对象
def get_workbook_sheet(path, list_file_name, file_name, sheet_name, read_only, data_only):
    route, manual_table_name = get_file_path(path, list_file_name, file_name)
    workbook = load_workbook(route, read_only=read_only, data_only=data_only)
    sheet_name = get_sheet_name_by_workbook(workbook, sheet_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet, manual_table_name, workbook


# 获取文件路径
def get_file_path(path, file_list, name):
    manual_table_name = ""
    for file_name in file_list:
        if str(file_name).count(name) > 0:
            manual_table_name = file_name

    route = path + "\\" + manual_table_name
    return route, manual_table_name


# 根据包含的名字获取sheet名字
def get_sheet_name_by_workbook(workbook, name):
    # 查看所有工作表
    sheet_names = workbook.sheetnames
    # print("查看所有工作表", sheet_names)
    work_sheet_name = ""
    # 遍历sheet
    for i in sheet_names:
        if i.__contains__(name):
            work_sheet_name = i
    # print("输出工作sheet名字\t" + work_sheet_name)
    return work_sheet_name


def get_list_file_by_path(wx, path):
    # print("path的数据类型是:",type(path))path的数据类型是: <class 'str'>
    if len(path) == 0:
        prompt_box(wx, "提示", "未选择目录程序结束")
        return -1
    try:
        # 获取所有文件
        list_file_name = os.listdir(path)
        for file_name in list_file_name:
            if file_name.endswith(".xls"):
                prompt_box(wx, "错误", "请检查 " + file_name + " 文件格式是否正确,希望是.xlsx")
                return -1
        return list_file_name
    except OSError:
        prompt_box(wx, "提示", "路径不正确")
        return -1


# 创建提示框
def prompt_box(wx, title, news):
    # 创建提示对话框
    dlg = wx.MessageDialog(None, news, title, wx.OK)
    # 显示对话框
    dlg.ShowModal()
    # 关闭对话框
    dlg.Destroy()


# 设置进度条
def set_m_gauge_value(global_self, x):
    global_self.m_gauge_进度条.SetValue(x)


# 获取字符串里的数字
def get_numbers(s):
    numbers = ''
    for c in s:
        if c.isdigit():
            numbers += c
    return numbers
